import openpyxl
import numpy as np

book = openpyxl.load_workbook('./output/input.xlsx')

book_edit = openpyxl.Workbook()
sheet_edit = book_edit.active

sheet = book['sheet1']
firstCol = 2
firstRow = 2
nCols = 5
nRows = 2
data = np.array([[i.value for i in j] for j in sheet.iter_rows()])
data = data[1:]

row_qty = 0

# пропуски
amount_pass = 0
initial_pr_pass = 0

# отрицательные
amount_negative = 0
initial_pr_negative = 0
new_row_table = 1
for row in data:
    is_negative = 0
    row_qty += 1
    if row[3] is None:
        amount_pass += 1
    elif row[3] < 0:
        amount_negative += 1
        is_negative = 1
    if row[4] is None:
        initial_pr_pass += 1
    elif row[4] < 0:
        initial_pr_negative += 1
        is_negative = 1

    if (is_negative == 0):
        new_row_table += 1
        sheet_edit['A' + str(new_row_table)] = str(row[0])
        sheet_edit['B' + str(new_row_table)] = str(row[1])
        sheet_edit['C' + str(new_row_table)] = row[2]
        sheet_edit['D' + str(new_row_table)] = row[3]
        sheet_edit['E' + str(new_row_table)] = row[4]

print('Всего строк ', row_qty)
print('Пропусков Amount ', amount_pass)
print('Пропусков initial ', initial_pr_pass)
print('Отрицательных Amount ', amount_negative)
print('Отрицательных initial ', initial_pr_negative)
negative = (amount_negative + initial_pr_negative) / row_qty
all_pass = (amount_pass + initial_pr_pass) / row_qty
print('Доля пропуско ', all_pass)
print('Доля отрицательных ', negative)

book_edit.save('./output/out.xlsx')
exit()
